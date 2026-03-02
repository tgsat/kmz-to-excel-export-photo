# Deskripsi :
# - Script ini membaca file CSV yang berisi nama file foto, memprosesnya untuk mengekstrak informasi seperti globalid dan nomor foto, lalu menyimpan hasilnya ke dalam file Excel dengan sheet terpisah untuk setiap nomor foto. Jika ada kolom "file_link", script ini juga akan mengaktifkan hyperlink di kolom tersebut.

# Catatan :
# - Pastikan file CSV input memiliki format yang sesuai dengan pola yang ditentukan dalam regex
# - support idpel_photo_index & idpel_foto_index format png & jpg.

import csv
import re
import os
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# === Konfigurasi ===
output_folder = r"D:\KANTOR\MAPPING\FECTH_GDRIVE\SSTB\ALLIN MAPMARKER\APP\27022026"
local_csv = os.path.join(output_folder, "app-27022026.csv")
excel_output = os.path.join(output_folder, "app-27022026.xlsx")

os.makedirs(output_folder, exist_ok=True)

pattern = re.compile(
    r'^{?([0-9a-fA-F-]{8,36})}?_(foto|photo)[_\-]?(\d+)\.(jpg|png)$',
    re.IGNORECASE
)

output_data = defaultdict(list)

if not os.path.exists(local_csv):
    raise FileNotFoundError(f"❌ File input tidak ditemukan: {local_csv}")

with open(local_csv, mode='r', newline='', encoding='utf-8') as infile:
    reader = csv.reader(infile)
    header = next(reader, None)
    for row in reader:
        if not row or len(row) < 1:
            continue

        raw_filename = row[0].strip()
        match = pattern.match(raw_filename)
        if match:
            globalid = match.group(1).replace('_', '-').upper()
            photo_number = match.group(3)
            ext = match.group(4).lower()

            label = match.group(2).lower()
            formatted_filename = f"{globalid}_{label}_{photo_number}.{ext}"

            new_row = [f"{globalid}", formatted_filename] + row[1:]
            output_data[photo_number].append(new_row)
        else:
            print(f"⚠️ Tidak cocok: {raw_filename}")

if output_data:
    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        for number, rows in sorted(output_data.items(), key=lambda x: int(x[0])):
            df = pd.DataFrame(rows, columns=['idpelanggan', 'filename'] + (header[1:] if header else []))
            sheet_name = f"photo_{number}"
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Sheet '{sheet_name}' ditambahkan ({len(rows)} baris).")

    wb = load_workbook(excel_output)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = [cell.value for cell in ws[1]]
        if "file_link" in [h.lower() for h in header_row]:
            col_index = [h.lower() for h in header_row].index("file_link") + 1
            col_letter = get_column_letter(col_index)

            for cell in ws[col_letter]:
                if cell.row == 1:
                    continue
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                    link = cell.value
                    cell.hyperlink = link
                    cell.font = Font(color="0000EE", underline="single")

            print(f"🔗 Hyperlink aktif diterapkan di sheet '{sheet_name}' (kolom file_link).")

    wb.save(excel_output)
    print(f"\n🎉 Semua data berhasil disimpan ke '{excel_output}' dengan hyperlink aktif.")
else:
    print("⚠️ Tidak ada data yang cocok dengan pola foto.")
